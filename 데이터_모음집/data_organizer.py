"""
데이터 분류 수집 프로그램
- 사용자가 분류 박스 이름을 지정
- 박스 안에 들어갈 데이터를 입력
- 엑셀 파일에 분류 박스별로 시트를 나누어 시간 오름차순 정렬하여 저장
- 작업 취소(되돌리기) 기능 지원
- 숫자 데이터 확률 계산 기능
"""

import os
import sys
from datetime import datetime
from collections import Counter

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("openpyxl 패키지가 필요합니다. 설치 중...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# 엑셀 파일 경로
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "분류_데이터.xlsx")

# 작업 히스토리 (되돌리기용)
history = []
MAX_HISTORY = 20


# ─────────────────────────────────────────────
# UI 유틸리티 (토스 스타일 미니멀 디자인)
# ─────────────────────────────────────────────

def clear_screen():
    os.system("cls" if os.name == "nt" else "clear")


def print_header():
    """토스 스타일 깔끔한 헤더"""
    print()
    print("  ┌────────────────────────────────────────────┐")
    print("  │                                            │")
    print("  │        📊  데이터 분류 수집기              │")
    print("  │                                            │")
    print("  │   분류 박스를 만들고, 데이터를 정리하세요  │")
    print("  │                                            │")
    print("  └────────────────────────────────────────────┘")
    print()


def print_divider():
    print("  ─ ─ ─ ─ ─ ─ ─ ─ ─ ─ ─ ─ ─ ─ ─ ─ ─ ─ ─ ─ ─")


def print_menu():
    """토스 스타일 메뉴"""
    print()
    print("  무엇을 할까요?")
    print()
    print("    1   데이터 입력")
    print("    2   분류 박스 목록")
    print("    3   박스 데이터 조회")
    print("    4   데이터 삭제")
    print("    5   확률 계산")
    print("    6   되돌리기")
    print("    7   박스 삭제")
    print("    8   저장 후 종료")
    print()


def print_success(msg):
    print(f"\n  ✓ {msg}")


def print_error(msg):
    print(f"\n  ✗ {msg}")


def print_info(msg):
    print(f"\n  ℹ {msg}")


def print_box_list(wb):
    """박스 목록을 깔끔하게 표시"""
    if not wb.sheetnames:
        print_info("등록된 분류 박스가 없습니다.")
        return

    print()
    print("  📦 분류 박스")
    print_divider()
    for i, name in enumerate(wb.sheetnames, 1):
        ws = wb[name]
        count = ws.max_row - 1 if ws.max_row > 1 else 0
        print(f"    {i}.  {name}  ─  {count}건")
    print_divider()


def print_box_data(wb, box_name):
    """박스 데이터를 깔끔하게 표시"""
    if box_name not in wb.sheetnames:
        print_error(f"'{box_name}' 박스가 존재하지 않습니다.")
        return

    ws = wb[box_name]
    if ws.max_row <= 1:
        print_info(f"[{box_name}] 박스에 데이터가 없습니다.")
        return

    print()
    print(f"  📋 {box_name}")
    print_divider()
    print(f"    {'#':<5} {'데이터':<30} {'입력 시간'}")
    print(f"    {'─'*5} {'─'*30} {'─'*19}")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] is not None:
            print(f"    {row[0]:<5} {str(row[1]):<30} {row[2]}")
    print_divider()


# ─────────────────────────────────────────────
# 핵심 로직
# ─────────────────────────────────────────────

def save_snapshot(wb, description):
    """현재 상태를 히스토리에 저장"""
    snapshot = {"description": description, "sheets": {}}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        snapshot["sheets"][sheet_name] = rows

    history.append(snapshot)
    if len(history) > MAX_HISTORY:
        history.pop(0)


def restore_snapshot(wb, snapshot):
    """스냅샷으로 워크북 복원"""
    for name in wb.sheetnames:
        del wb[name]

    for sheet_name, rows in snapshot["sheets"].items():
        ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)
        if ws.max_row >= 1:
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 50
            ws.column_dimensions["C"].width = 22
            style_header(ws)


def load_workbook():
    """기존 엑셀 파일을 불러오거나 새로 생성"""
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
    return wb


def style_header(ws):
    """시트 헤더에 스타일 적용"""
    header_fill = PatternFill(start_color="1B1B1B", end_color="1B1B1B", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for cell in ws[1]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border


def get_or_create_sheet(wb, box_name):
    """분류 박스에 해당하는 시트를 가져오거나 새로 생성"""
    if box_name in wb.sheetnames:
        return wb[box_name]
    else:
        ws = wb.create_sheet(title=box_name)
        ws.append(["번호", "데이터", "입력 시간"])
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 22
        style_header(ws)
        return ws


def add_data_to_sheet(ws, data):
    """시트에 데이터 추가 후 시간 오름차순 정렬"""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] is not None:
            rows.append((row[1], row[2]))

    rows.append((data, now))
    rows.sort(key=lambda x: x[1])

    ws.delete_rows(2, ws.max_row)

    for idx, (d, t) in enumerate(rows, start=1):
        ws.append([idx, d, t])


def delete_data_from_sheet(ws, data_index):
    """시트에서 특정 번호의 데이터 삭제"""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] is not None:
            rows.append((row[1], row[2]))

    if data_index < 1 or data_index > len(rows):
        return False

    rows.pop(data_index - 1)
    ws.delete_rows(2, ws.max_row)

    rows.sort(key=lambda x: x[1])
    for idx, (d, t) in enumerate(rows, start=1):
        ws.append([idx, d, t])

    return True


def calculate_probability(wb, box_name, target_value):
    """
    박스 내 숫자 데이터의 확률을 계산합니다.
    target_value가 전체 데이터 중 몇 %를 차지하는지 반환합니다.
    """
    if box_name not in wb.sheetnames:
        return None, "박스가 존재하지 않습니다."

    ws = wb[box_name]
    if ws.max_row <= 1:
        return None, "박스에 데이터가 없습니다."

    # 모든 데이터를 숫자로 변환 시도
    all_values = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] is not None:
            try:
                val = float(row[1])
                all_values.append(val)
            except (ValueError, TypeError):
                # 숫자가 아닌 데이터는 건너뜀
                pass

    if not all_values:
        return None, "박스에 숫자 데이터가 없습니다."

    total = len(all_values)
    count = all_values.count(target_value)
    probability = (count / total) * 100

    return {
        "target": target_value,
        "count": count,
        "total": total,
        "probability": probability,
        "distribution": Counter(all_values),
    }, None


def show_probability_result(result):
    """확률 계산 결과를 토스 스타일로 표시"""
    print()
    print("  ┌────────────────────────────────────────────┐")
    print("  │                                            │")
    print(f"  │   🎯 {result['target']}의 확률                         │")
    print("  │                                            │")

    # 확률을 크게 표시
    prob_str = f"{result['probability']:.1f}%"
    print(f"  │          {prob_str:>10}                       │")
    print("  │                                            │")
    print(f"  │   전체 {result['total']}건 중 {result['count']}건            │")
    print("  │                                            │")
    print("  └────────────────────────────────────────────┘")

    # 전체 분포도 표시
    print()
    print("  📊 전체 분포")
    print_divider()

    distribution = result["distribution"]
    total = result["total"]

    # 값 기준 정렬
    for value in sorted(distribution.keys()):
        cnt = distribution[value]
        pct = (cnt / total) * 100
        bar_len = int(pct / 5)  # 최대 20칸
        bar = "█" * bar_len + "░" * (20 - bar_len)

        # 정수로 표시 가능하면 정수로
        val_display = int(value) if value == int(value) else value
        print(f"    {val_display:<8} {bar}  {pct:.1f}% ({cnt}건)")

    print_divider()


# ─────────────────────────────────────────────
# 메인 루프
# ─────────────────────────────────────────────

def main():
    clear_screen()
    print_header()
    print(f"  저장 위치: {EXCEL_FILE}")

    wb = load_workbook()

    while True:
        print_menu()
        choice = input("  선택 → ").strip()

        if choice == "1":
            # 데이터 입력
            print_box_list(wb)
            print()
            print("  새 박스 이름을 입력하면 자동 생성됩니다.")
            box_name = input("  📦 박스 이름 (취소: 빈칸) → ").strip()

            if not box_name:
                print_info("메뉴로 돌아갑니다.")
                continue
            if len(box_name) > 31:
                print_error("박스 이름은 31자 이내로 입력해주세요.")
                continue

            save_snapshot(wb, f"[{box_name}] 데이터 입력 전")
            ws = get_or_create_sheet(wb, box_name)

            print()
            print(f"  📝 [{box_name}]에 데이터를 입력하세요.")
            print("     빈 줄 → 입력 종료  |  '취소' → 전체 취소")
            print()

            count = 0
            cancelled = False
            while True:
                data = input(f"    {count + 1}번째 데이터 → ").strip()
                if not data:
                    break
                if data == "취소":
                    cancelled = True
                    break
                add_data_to_sheet(ws, data)
                count += 1

            if cancelled:
                if history:
                    snapshot = history.pop()
                    restore_snapshot(wb, snapshot)
                print_info("입력이 취소되었습니다.")
            elif count > 0:
                print_success(f"[{box_name}]에 {count}건 추가 완료")
                wb.save(EXCEL_FILE)
                print("  💾 자동 저장됨")
            else:
                if history and history[-1]["description"] == f"[{box_name}] 데이터 입력 전":
                    history.pop()
                print_info("입력된 데이터가 없습니다.")

        elif choice == "2":
            print_box_list(wb)

        elif choice == "3":
            print_box_list(wb)
            if wb.sheetnames:
                box_name = input("\n  조회할 박스 이름 (취소: 빈칸) → ").strip()
                if not box_name:
                    continue
                print_box_data(wb, box_name)

        elif choice == "4":
            # 데이터 삭제
            print_box_list(wb)
            if not wb.sheetnames:
                continue

            box_name = input("\n  삭제할 데이터가 있는 박스 (취소: 빈칸) → ").strip()
            if not box_name:
                continue
            if box_name not in wb.sheetnames:
                print_error(f"'{box_name}' 박스가 존재하지 않습니다.")
                continue

            print_box_data(wb, box_name)
            ws = wb[box_name]
            if ws.max_row <= 1:
                continue

            try:
                idx = input("  삭제할 번호 (취소: 빈칸) → ").strip()
                if not idx:
                    continue
                idx = int(idx)
            except ValueError:
                print_error("숫자를 입력해주세요.")
                continue

            save_snapshot(wb, f"[{box_name}] {idx}번 삭제 전")

            if delete_data_from_sheet(ws, idx):
                print_success(f"{idx}번 데이터 삭제 완료")
                wb.save(EXCEL_FILE)
                print("  💾 자동 저장됨")
            else:
                history.pop()
                print_error("잘못된 번호입니다.")

        elif choice == "5":
            # 확률 계산
            print_box_list(wb)
            if not wb.sheetnames:
                continue

            box_name = input("\n  확률을 계산할 박스 이름 (취소: 빈칸) → ").strip()
            if not box_name:
                continue
            if box_name not in wb.sheetnames:
                print_error(f"'{box_name}' 박스가 존재하지 않습니다.")
                continue

            # 먼저 현재 데이터 보여주기
            print_box_data(wb, box_name)

            print()
            target_input = input("  🎯 확률을 알고 싶은 값 (취소: 빈칸) → ").strip()
            if not target_input:
                continue

            try:
                target_value = float(target_input)
            except ValueError:
                print_error("숫자를 입력해주세요.")
                continue

            result, error = calculate_probability(wb, box_name, target_value)
            if error:
                print_error(error)
            else:
                show_probability_result(result)

        elif choice == "6":
            # 되돌리기
            if not history:
                print_info("되돌릴 작업이 없습니다.")
                continue

            print()
            print("  🕐 최근 작업")
            print_divider()
            for i, h in enumerate(reversed(history), 1):
                print(f"    {i}.  {h['description']}")
                if i >= 5:
                    break
            print_divider()

            confirm = input("\n  가장 최근 작업을 되돌릴까요? (y/n) → ").strip().lower()
            if confirm == "y":
                snapshot = history.pop()
                restore_snapshot(wb, snapshot)
                wb.save(EXCEL_FILE)
                print_success(f"되돌리기 완료: {snapshot['description']}")
                print("  💾 저장됨")
            else:
                print_info("취소했습니다.")

        elif choice == "7":
            # 박스 삭제
            print_box_list(wb)
            if not wb.sheetnames:
                continue

            box_name = input("\n  삭제할 박스 이름 (취소: 빈칸) → ").strip()
            if not box_name:
                continue
            if box_name not in wb.sheetnames:
                print_error(f"'{box_name}' 박스가 존재하지 않습니다.")
                continue

            confirm = input(f"  ⚠️  [{box_name}] 박스를 정말 삭제할까요? (y/n) → ").strip().lower()
            if confirm == "y":
                save_snapshot(wb, f"[{box_name}] 박스 삭제 전")
                del wb[box_name]
                wb.save(EXCEL_FILE)
                print_success(f"[{box_name}] 박스 삭제 완료")
                print("  💾 저장됨")
            else:
                print_info("취소했습니다.")

        elif choice == "8":
            wb.save(EXCEL_FILE)
            print()
            print_success(f"저장 완료: {EXCEL_FILE}")
            print()
            print("  👋 다음에 또 만나요!")
            print()
            break

        else:
            print_error("1~8 중에서 선택해주세요.")


if __name__ == "__main__":
    main()
